# -*- coding: utf-8 -*-
"""
Created on Wed Oct 12 10:24:15 2022
@author: SEN2613

PORTFOLIO VERSION
-----------------
This version has been anonymized for confidentiality purposes.

Sensitive information from the original professional environment,
including network paths and workbook passwords, has been replaced
with fictitious portfolio values.

The original automation logic and technical approach have been preserved.
"""

###############################################################################
# PACKAGES
###############################################################################

import pandas as pd
import re
import time
import os
from openpyxl import load_workbook
import shutil
from openpyxl.worksheet.datavalidation import DataValidation
import warnings

# Ignore non-critical Excel warnings during automated workbook processing
warnings.simplefilter(action='ignore', category=UserWarning)


###############################################################################
# FONCTIONS
###############################################################################

def Ouvrir_Excel_File(excel_dest):
    """
    Opens the Excel workbook, makes hidden worksheets visible,
    unlocks the workbook structure and identifies the Base worksheet.
    """

    time.sleep(5)

    try:
        wb = load_workbook(excel_dest)
    except FileNotFoundError:
        raise FileNotFoundError(f"[ERREUR] Fichier Excel introuvable : {excel_dest}")
    except PermissionError:
        raise PermissionError(f"[ERREUR] Fichier Excel verrouillé (ouvert par un autre utilisateur) : {excel_dest}")

    time.sleep(5)

    # Portfolio placeholder used instead of the original workbook password
    wb.security.workbookPassword = 'xxxxxxxxx'
    wb.security.lockStructure = False

    # Make hidden worksheets visible before processing
    for i in wb.worksheets:
        if i.sheet_state == "hidden":
            i.sheet_state = "visible"

    # Identify the worksheet containing the Base data
    feuille = [s for s in wb.sheetnames if "Base" in s]

    if not feuille:
        raise ValueError(f"[ERREUR] Aucun onglet 'Base' trouvé dans : {excel_dest}")

    wb.active = wb[feuille[0]]
    wb.save(excel_dest)

    return feuille[0]


def Masquer_Onglet(fichierScorecards):
    """
    Hides technical worksheets and keeps the user-facing dashboard visible.
    """

    try:
        wb = load_workbook(fichierScorecards)
    except FileNotFoundError:
        raise FileNotFoundError(f"[ERREUR] Fichier introuvable pour masquage : {fichierScorecards}")

    # Identify the user-facing worksheets
    onglet1 = [s for s in wb.sheetnames if "Dashboard" in s]
    onglet2 = [s for s in wb.sheetnames if "Barème" in s]

    if not onglet1:
        raise ValueError(f"[ERREUR] Onglet 'Dashboard' introuvable dans : {fichierScorecards}")

    if not onglet2:
        raise ValueError(f"[ERREUR] Onglet 'Barème' introuvable dans : {fichierScorecards}")

    onglets = list(wb.sheetnames)

    onglets.remove(onglet1[0])
    onglets.remove(onglet2[0])

    # Restaurant Scorecards also keep the Scorecard worksheet visible
    if "Restaurants" in fichierScorecards:
        onglet3 = [s for s in wb.sheetnames if "Scorecard" in s]

        if not onglet3:
            raise ValueError(f"[ERREUR] Onglet 'Scorecard' introuvable dans : {fichierScorecards}")

        onglets.remove(onglet3[0])

    # Hide all remaining technical worksheets
    for onglet in onglets:
        wb[onglet].sheet_state = 'hidden'

    # Portfolio placeholder used instead of the original workbook password
    wb.security.workbookPassword = 'xxxxxxxxx'
    wb.security.lockStructure = True

    # Keep the Dashboard as the active worksheet
    wb.active = wb[onglet1[0]]

    wb.save(fichierScorecards)


def Supprimer_Lignes(excel_dest, feuille):
    """
    Clears existing data from the Excel worksheet before importing
    the latest reporting data.
    """

    try:
        wb = load_workbook(excel_dest)
    except FileNotFoundError:
        raise FileNotFoundError(f"[ERREUR] Fichier introuvable pour suppression de lignes : {excel_dest}")

    if feuille not in wb.sheetnames:
        raise ValueError(f"[ERREUR] Onglet '{feuille}' introuvable dans : {excel_dest}")

    sheet = wb[feuille]

    # Delete existing data while keeping the worksheet structure
    sheet.delete_rows(2, sheet.max_row + 1)

    wb.save(excel_dest)


def iteration_cellule(sheet, plage, valeurs):
    """
    Writes imported values into predefined Excel cells.
    """

    for i, colonne in enumerate(plage):
        sheet.cell(row=4, column=colonne).value = valeurs.iloc[i]


def Titre(ws2, mois, typescor):
    """
    Updates the report title according to the reporting month
    and Scorecard type.
    """

    for x in range(2, 3):
        ws2.merge_cells(
            start_row=x,
            start_column=3,
            end_row=x,
            end_column=4
        )

    moislettre = re.sub(r'[\W\d_]+', '', mois)

    if len(typescor) > 0:
        ws2['C2'] = moislettre + typescor
    else:
        ws2['C2'] = moislettre


def Copier_Coller_Entete(fichierScorecards, fichierImporte, typescor):
    """
    Updates the Dashboard header using data from the imported CSV file.
    """

    try:
        wb1 = load_workbook(fichierScorecards)
    except FileNotFoundError:
        raise FileNotFoundError(f"[ERREUR] Fichier introuvable pour copie entête : {fichierScorecards}")

    # Identify the Dashboard worksheet
    feuille = [s for s in wb1.sheetnames if "Dashboard" in s]

    if not feuille:
        raise ValueError(f"[ERREUR] Onglet 'Dashboard' introuvable dans : {fichierScorecards}")

    ws2 = wb1[feuille[0]]

    # Retrieve the reporting month from the folder structure
    recupmois = fichierScorecards.split('\\')
    mois = recupmois[10]

    Titre(ws2, mois, typescor)

    # Update the reporting header
    ws2.cell(row=4, column=5).value = fichierImporte.iloc[0, 0]

    # Clients
    valeur = fichierImporte.iloc[0, 1:7]
    iteration_cellule(ws2, range(7,13), valeur)

    # Propreté
    valeur = fichierImporte.iloc[0, 7:9]
    iteration_cellule(ws2, range(14,16), valeur)

    # Goût
    valeur = fichierImporte.iloc[0, 9:11]
    iteration_cellule(ws2, range(17,19), valeur)

    # Service Speed
    valeur = fichierImporte.iloc[0, 11:13]
    iteration_cellule(ws2, range(20,22), valeur)

    # Formation
    valeur = fichierImporte.iloc[0, 13:19]
    iteration_cellule(ws2, range(23,29), valeur)

    # ROCC
    valeur = fichierImporte.iloc[0, 19:23]
    iteration_cellule(ws2, range(30,34), valeur)

    wb1.save(fichierScorecards)


###############################################################################
# LISTE DÉROULANTE SCORECARD RESTAURANT
###############################################################################

def listederoulante(excel_dest, feuille, listerestaurant):
    """
    Creates an Excel dropdown list containing the available restaurants.
    """

    try:
        wb = load_workbook(excel_dest)
    except FileNotFoundError:
        raise FileNotFoundError(f"[ERREUR] Fichier introuvable pour liste déroulante : {excel_dest}")

    if feuille not in wb.sheetnames:
        raise ValueError(f"[ERREUR] Onglet '{feuille}' introuvable dans : {excel_dest}")

    sheet = wb[feuille]

    if not isinstance(listerestaurant, list):
        listerestaurant = list(map(str, listerestaurant))

    # Temporary reference column used to store dropdown values
    col_ref = "Z"

    for i, val in enumerate(listerestaurant, start=1):
        sheet[f"{col_ref}{i}"] = val

    plage = f"{col_ref}1:{col_ref}{len(listerestaurant)}"

    # Create Excel data validation
    data_validation = DataValidation(
        type="list",
        formula1=f"={plage}",
        allow_blank=True
    )

    sheet.add_data_validation(data_validation)

    # Apply dropdown to the restaurant selection cell
    data_validation.add(sheet["B1"])

    wb.save(excel_dest)


###############################################################################
# CRÉATION NOUVEAU RÉPERTOIRE
###############################################################################

def Creation_Repertoire_Fichier(dossierprecedent, dossieracree, fichiers, moisprecedent, moisatraite, param):
    """
    Creates the reporting folder and copies the previous reporting files
    as a starting point for the new reporting period.
    """

    if len([s for s in fichiers if dossieracree in s]) == 0:

        # Create the new reporting directory
        os.mkdir(dossieracree)

        print(f"[INFO] Dossier créé : {dossieracree}")

        # Copy files from the previous reporting period
        for fichierscopies in os.listdir(dossierprecedent):

            src_file = os.path.join(dossierprecedent, fichierscopies)

            try:
                shutil.copy(src_file, dossieracree)
                print(f"[INFO] Fichier copié : {fichierscopies}")

            except PermissionError:
                print(f"[ERREUR] Impossible de copier (fichier verrouillé) : {fichierscopies}")
                continue

            except shutil.Error as e:
                print(f"[ERREUR] Copie échouée pour {fichierscopies} : {e}")
                continue

            # Update the reporting period and year in Scorecard filenames
            if fichierscopies.find("Scorecard Restaurants") != -1 or fichierscopies.find("Scorecard Franchisés") != -1:

                nouveaunom = fichierscopies.replace(
                    "M" + str(moisprecedent),
                    moisatraite
                ).replace(
                    fichierscopies[len(fichierscopies) - 9: len(fichierscopies) - 5],
                    str(param.iloc[0, 1])
                )

                dstfichier = os.path.join(dossieracree, fichierscopies)
                nouveaudstfichier = os.path.join(dossieracree, nouveaunom)

                try:
                    os.rename(dstfichier, nouveaudstfichier)

                except Exception as e:
                    print(f"[ERREUR] Renommage échoué pour {fichierscopies} : {e}")


###############################################################################
# BALAYAGE DES FICHIERS
###############################################################################

def balayage_fichiers():
    """
    Identifies the previous and current reporting directories
    based on the reporting calendar and parameters.
    """

    try:

        # Portfolio path replacing the original internal network location
        calendrier = pd.read_excel(
            r"\\PORTFOLIO-FILESRV01\Portfolio\Reporting\Automation\Parametrage.xlsx",
            sheet_name="Calendrier"
        ).set_index('mnum').T.to_dict('list')

        param = pd.read_excel(
            r"\\PORTFOLIO-FILESRV01\Portfolio\Reporting\Automation\Parametrage.xlsx",
            sheet_name="Param",
            header=None
        )

    except FileNotFoundError:
        raise FileNotFoundError(
            "[ERREUR] Fichier Parametrage.xlsx introuvable sur le réseau."
        )

    moisatraite = param.iloc[0, 0]
    anneeatraite = param.iloc[0, 1]

    # Portfolio reporting directory
    chemin = f"\\\\PORTFOLIO-FILESRV01\\Portfolio\\Reporting\\Scorecards\\Scorecard {anneeatraite}\\"

    fichiers = list([x[0] for x in os.walk(chemin)])

    # Identify the previous reporting month
    if len(moisatraite) == 3:

        moisprecedent = int(moisatraite[1:3]) - 1

        dossierprecedent = chemin + calendrier[
            "M" + str(moisprecedent)
        ][0]

        dossieracree = chemin + calendrier[
            moisatraite
        ][0]

    elif int(moisatraite[1]) == 1:

        moisprecedent = 12

        cheminprecedent = chemin.replace(
            chemin[len(chemin) - 5: len(chemin) - 1],
            str(param.iloc[0, 1])
        )

        dossierprecedent = cheminprecedent + calendrier[
            "M" + str(moisprecedent)
        ]

        dossieracree = chemin + calendrier[
            moisatraite
        ][0]

    else:

        moisprecedent = int(moisatraite[1]) - 1

        dossierprecedent = chemin + calendrier[
            "M" + str(moisprecedent)
        ][0]

        dossieracree = chemin + calendrier[
            moisatraite
        ][0]

    print(f"[INFO] Dossier cible : {dossieracree}")

    # Remove the existing reporting folder before recreating it
    if dossieracree in fichiers:

        shutil.rmtree(
            dossieracree,
            ignore_errors=True
        )

        print("[INFO] Dossier existant supprimé")

        time.sleep(7)

        # Refresh directory list after deletion
        fichiers = list([x[0] for x in os.walk(chemin)])

    print("[INFO] Création du dossier en cours...")

    os.listdir(chemin)

    Creation_Repertoire_Fichier(
        dossierprecedent,
        dossieracree,
        fichiers,
        moisprecedent,
        moisatraite,
        param
    )

    return dossieracree


###############################################################################
# CONTEXT : DEV / PROD
###############################################################################

# Production environment
context = "PROD"

# Portfolio version of the configuration file path
PARAMETRAGE_PATH = (
    r"\\PORTFOLIO-FILESRV01\Portfolio\Reporting\Automation\Parametrage.xlsx"
)

try:

    # Load Scorecard file configuration
    fichiersScorecard = pd.read_excel(
        PARAMETRAGE_PATH,
        sheet_name="Fichiers_Scorecards"
    ).set_index('index').T.to_dict('list')

    # Load imported file configuration
    fichiersImportes = pd.read_excel(
        PARAMETRAGE_PATH,
        sheet_name="Fichiers_Importes"
    ).set_index('index').T.to_dict('list')

except FileNotFoundError:

    raise FileNotFoundError(
        f"[ERREUR] Impossible de lire le fichier de paramétrage : {PARAMETRAGE_PATH}"
    )


# Development environment
if context == 'DEV':

    repertoiresScorecard = (
        r"\\PORTFOLIO-FILESRV01\Portfolio\Reporting\Automation\Development\Test"
        + '\\'
    )

    repertoireImporte = (
        r"\\PORTFOLIO-FILESRV01\Portfolio\Reporting\Automation"
        + '\\'
    )

else:

    # Production reporting directory
    repertoiresScorecard = balayage_fichiers() + '\\'

    # Directory containing imported CSV files
    repertoireImporte = (
        r"\\PORTFOLIO-FILESRV01\Portfolio\Reporting\Automation"
        + '\\'
    )


###############################################################################
# BASE DONNÉES RESTAURANTS / FRANCHISÉS
###############################################################################

# Column used to rank the Scorecards
COLONNE_TRI = "Note Totale"


# Process each configured Scorecard
for index in fichiersScorecard.keys():

    try:

        # Build the full paths to the Scorecard and source CSV
        fichierScorecards = (
            repertoiresScorecard
            + fichiersScorecard.get(index)[0]
        )

        filename = (
            repertoireImporte
            + fichiersImportes.get(index)[0]
        )

        # Open and prepare the Excel Scorecard
        sheet = Ouvrir_Excel_File(fichierScorecards)

        # Remove previous data before importing the latest data
        Supprimer_Lignes(
            fichierScorecards,
            sheet
        )

        try:

            # Load the latest CSV data
            importer = pd.read_csv(
                filename,
                sep=','
            )

        except FileNotFoundError:

            print(
                f"[ERREUR] CSV introuvable pour {index} : {filename}"
            )

            continue

        except pd.errors.ParserError:

            print(
                f"[ERREUR] CSV mal formé pour {index} : {filename}"
            )

            continue

        # Validate the main ranking column
        if COLONNE_TRI not in importer.columns:

            print(
                f"[ERREUR] Colonne '{COLONNE_TRI}' absente dans {filename}"
            )

            print(
                f"         Colonnes disponibles : {list(importer.columns)}"
            )

            continue

        # Additional processing for restaurant Scorecards
        if "restaurants" in filename:

            colonnes_requises = [
                'Consultant',
                'Franchisé',
                'Restaurant'
            ]

            colonnes_manquantes = [
                c for c in colonnes_requises
                if c not in importer.columns
            ]

            if colonnes_manquantes:

                print(
                    f"[ERREUR] Colonnes manquantes dans {filename} : "
                    f"{colonnes_manquantes}"
                )

                continue

            # Remove incomplete records
            importer.dropna(
                subset=colonnes_requises,
                thresh=3
            )

            # Prepare the restaurant list for the dropdown
            liste_restaurant = importer["Restaurant"].str.upper()

            liste_restaurant.sort_values(
                ascending=True,
                inplace=True
            )

            # Add the restaurant selection dropdown
            listederoulante(
                fichierScorecards,
                "Scorecard",
                liste_restaurant
            )

        # Sort the reporting data by total score
        importer.sort_values(
            COLONNE_TRI,
            ascending=False,
            inplace=True
        )

        # Write the processed data into the existing Excel workbook
        with pd.ExcelWriter(
            fichierScorecards,
            engine="openpyxl",
            mode="a",
            if_sheet_exists='overlay'
        ) as writer:

            importer.to_excel(
                writer,
                sheet_name=sheet,
                header=None,
                startrow=1,
                index=False
            )

        print(f"[OK] Traitement terminé pour : {index}")

    except Exception as e:

        print(
            f"[ERREUR] Traitement échoué pour {index} : "
            f"{type(e).__name__} - {e}"
        )

        continue


###############################################################################
# ENTÊTES
###############################################################################

# Process Scorecard headers and reporting information
for index in list(fichiersImportes.keys()):

    if index not in list(fichiersScorecard.keys()):

        try:

            # Determine the Scorecard type
            if index[0:3] == "sem":

                cibles = [
                    fichiersScorecard['sfm'],
                    fichiersScorecard['srm']
                ]

                typescor = ""

            else:

                cibles = [
                    fichiersScorecard['sfytd'],
                    fichiersScorecard['srytd']
                ]

                typescor = "YTD"

            # Retrieve the CSV used for the header
            fichier_csv = (
                repertoireImporte
                + fichiersImportes.get(index)[0]
            )

            try:

                fichierImporte = pd.read_csv(
                    fichier_csv,
                    sep=','
                )

            except FileNotFoundError:

                print(
                    f"[ERREUR] CSV entête introuvable pour "
                    f"{index} : {fichier_csv}"
                )

                continue

            except pd.errors.ParserError:

                print(
                    f"[ERREUR] CSV entête mal formé pour "
                    f"{index} : {fichier_csv}"
                )

                continue

            # Apply the header information to each target Scorecard
            for fichier in cibles:

                fichierScorecards = (
                    repertoiresScorecard
                    + fichier[0]
                )

                Copier_Coller_Entete(
                    fichierScorecards,
                    fichierImporte,
                    typescor
                )

                # Hide technical worksheets after processing
                Masquer_Onglet(
                    fichierScorecards
                )

                print(
                    f"[OK] Entête appliquée : {fichier[0]}"
                )

        except Exception as e:

            print(
                f"[ERREUR] Entête échouée pour {index} : "
                f"{type(e).__name__} - {e}"
            )

            continue